"""Excel (XLSX) document processor for PII redaction."""

import os
from typing import List, Optional, Callable, Tuple
from copy import copy

from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell

from .base import BaseProcessor, BaseRestorer, ProcessingResult
from ..detectors.pii_detector import PIIDetector
from ..detectors.patterns import PIIMatch
from ..storage.mapping_store import MappingStore


class XlsxProcessor(BaseProcessor):
    """
    Process Excel (XLSX) files to redact PII.

    Processes all sheets, cell by cell, preserving formatting.
    """

    SUPPORTED_EXTENSIONS = ['.xlsx', '.xlsm']

    def __init__(
        self,
        detector: PIIDetector,
        mapping_store: MappingStore,
        interactive: bool = True,
        user_callback: Optional[Callable[[PIIMatch], Optional[bool]]] = None,
    ):
        super().__init__(detector, mapping_store, interactive, user_callback)

    def extract_text(self, input_path: str) -> str:
        """Extract all text from an Excel file."""
        wb = load_workbook(input_path, data_only=True)
        all_text = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"--- Sheet: {sheet_name} ---")

            for row in ws.iter_rows():
                row_texts = []
                for cell in row:
                    if cell.value is not None:
                        row_texts.append(str(cell.value))
                if row_texts:
                    all_text.append("\t".join(row_texts))

        wb.close()
        return "\n".join(all_text)

    def process(
        self,
        input_path: str,
        output_path: Optional[str] = None,
    ) -> ProcessingResult:
        """
        Process an Excel file, redacting PII.

        Args:
            input_path: Path to input Excel file
            output_path: Path for output file. If None, creates <input>_redacted.xlsx

        Returns:
            ProcessingResult with statistics
        """
        if output_path is None:
            base, ext = os.path.splitext(input_path)
            output_path = f"{base}_redacted{ext}"

        # Load workbook (preserve formatting)
        wb = load_workbook(input_path)

        total_redactions = 0
        redaction_counts = {}
        uncertain_count = 0

        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None or not isinstance(cell.value, str):
                        continue

                    text = str(cell.value)
                    if not text.strip():
                        continue

                    # Process cell text
                    processed_text, stats = self.process_text(text)

                    if stats['redacted'] > 0:
                        cell.value = processed_text
                        total_redactions += stats['redacted']
                        uncertain_count += stats['uncertain']

                        for pii_type, count in stats['by_type'].items():
                            redaction_counts[pii_type] = redaction_counts.get(pii_type, 0) + count

        # Save workbook
        wb.save(output_path)
        wb.close()

        return ProcessingResult(
            input_path=input_path,
            output_path=output_path,
            total_redactions=total_redactions,
            redactions_by_type=redaction_counts,
            uncertain_count=uncertain_count,
        )


class XlsxRestorer(BaseRestorer):
    """Restore redacted Excel files using mapping store."""

    def __init__(self, mapping_store: MappingStore):
        super().__init__(mapping_store)

    def restore(
        self,
        input_path: str,
        output_path: Optional[str] = None,
    ) -> Tuple[str, int]:
        """
        Restore a redacted Excel file.

        Args:
            input_path: Path to redacted Excel file
            output_path: Path for restored file

        Returns:
            Tuple of (output_path, restoration_count)
        """
        if output_path is None:
            base, ext = os.path.splitext(input_path)
            if base.endswith('_redacted'):
                base = base[:-9]
            output_path = f"{base}_restored{ext}"

        # Get restorations
        restorations = self.mapping_store.get_all_restorations()
        if not restorations:
            raise ValueError("No mappings available for restoration")

        # Load workbook
        wb = load_workbook(input_path)
        restoration_count = 0

        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    text = str(cell.value)

                    # Apply restorations
                    for placeholder, original in restorations.items():
                        if placeholder in text:
                            count = text.count(placeholder)
                            text = text.replace(placeholder, original)
                            restoration_count += count

                    if text != str(cell.value):
                        cell.value = text

        # Save workbook
        wb.save(output_path)
        wb.close()

        return output_path, restoration_count
